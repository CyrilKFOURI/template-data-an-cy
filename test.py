import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def power_category_summary(nova, output_file="Power_Category_Summary.xlsx"):

    # ==========================
    # Create summary table
    # ==========================
    counts = nova["POWER_CATEGORY"].value_counts(dropna=False)

    summary = pd.DataFrame({
        "Energy": counts.index.astype(str),
        "Share (%)": ((counts / counts.sum()) * 100).round(2),
        "Nb of Units": counts.values
    })

    total_row = pd.DataFrame({
        "Energy": ["Total"],
        "Share (%)": [100.00],
        "Nb of Units": [counts.sum()]
    })

    summary = pd.concat([summary, total_row], ignore_index=True)

    # ==========================
    # Create Pie Chart
    # ==========================
    fig, ax = plt.subplots(figsize=(7,7))

    wedges, texts, autotexts = ax.pie(
        counts,
        labels=counts.index.astype(str),
        autopct="%1.1f%%",
        startangle=90,
        pctdistance=0.80
    )

    centre_circle = plt.Circle((0,0),0.55,fc='white')
    fig.gca().add_artist(centre_circle)

    ax.set_title(
        "Distribution of Energy Categories - IN FLEET",
        fontsize=14,
        fontweight='bold'
    )

    plt.tight_layout()

    chart_file = "power_category_pie.png"
    plt.savefig(chart_file, dpi=300, bbox_inches="tight")
    plt.close()

    # ==========================
    # Export table to Excel
    # ==========================
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        summary.to_excel(
            writer,
            sheet_name="Energy Summary",
            index=False,
            startrow=1
        )

        workbook = writer.book
        ws = writer.sheets["Energy Summary"]

        ws["A1"] = "POWER CATEGORY SUMMARY"
        ws["A1"].font = Font(size=16, bold=True)

        # Style
        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        white_font = Font(color="FFFFFF", bold=True)

        thin = Side(style="thin")
        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        # Header formatting
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Table formatting
        for row in ws.iter_rows(min_row=3,
                                max_row=ws.max_row,
                                min_col=1,
                                max_col=3):

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # Highlight Total row
        total_row_excel = ws.max_row

        for cell in ws[total_row_excel]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D9EAD3",
                end_color="D9EAD3",
                fill_type="solid"
            )

        # Auto column width
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = length + 5

        # Insert chart
        img = Image(chart_file)
        img.width = 550
        img.height = 550

        ws.add_image(img, "E2")

    print(f"Excel generated : {output_file}")

    return summary